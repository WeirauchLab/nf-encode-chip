#!/usr/bin/env python

from argparse import ArgumentParser
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.cell import get_column_letter
import pandas as pd
import yaml

parser = ArgumentParser()
parser.add_argument("--config", default="sample_metrics.yaml", help="YAML file that defines columns in the output")
parser.add_argument("--motif", help="YAML file defining key motifs per group")
parser.add_argument("--byrow", dest="bycol", action="store_false", help="List samples in rows and metrics in columns")
parser.add_argument("--output", "-o", default="ChIP_metrics", help="Output filename, without an extenstion.  Default: ChIP_metrics")
parser.add_argument("multiqc_dir")
args = parser.parse_args()

with open(args.config) as f:
    metric_def = yaml.safe_load(f)

# More validationto do:
# 1) Validate structure/content of config file
# 2) Strip extension from output file (but don't strip path)

if args.motif:
    with open(args.motif) as f:
        key_motifs = yaml.safe_load(f)
else:
    key_motifs = {}

def read_metrics_file(root, filename, col_mapping, *, dtype=None, index_col='Sample', use='asis'):
    df = pd.read_table(
        root / filename,
        index_col=index_col,
        usecols=list(col_mapping)+[index_col],
        dtype=dtype
    ).rename(columns=col_mapping)
    if use == 'asis':
        return df
    if use == 'read1':
        return df.rename(index=lambda x: x.removesuffix('_1').removesuffix('_read1'))
    if use == 'read2':
        return df.rename(index=lambda x: x.removesuffix('_2').removesuffix('_read2'))
    raise ValueError(f"The use argument should be one of 'asis', 'read1', or 'read2' ({use} given)")

def collect_experiment(multiqc_path, motifs_of_interest):
    df = pd.concat(
        [ read_metrics_file(
            multiqc_path, 
            metric_def['definitions'][i]['file'], 
            metric_def['definitions'][i]['cols'],
            dtype=metric_def['definitions'][i].get('dtype', None),
            index_col=metric_def['definitions'][i].get('sample_col', 'Sample'),
            use=metric_def['definitions'][i].get('use_sample', 'asis'),
        ) for i in range(len(metric_def['definitions'])) ],
        axis=1,
        join='inner'
    ).sort_index().sort_values('Group', kind='stable')
    df = df[metric_def['order']]
    
    peak_info = pd.concat(
        [ get_encode_peak_stats(
            peakset,
            multiqc_path / f"multiqc_encode_{peakset.lower()}_reproducibility.txt"
        ) for peakset in ("IDR", "Overlap") ],
        axis=1
    ).merge(
        parse_motif_results(
            pd.read_table(
                multiqc_path / "multiqc_homer_findmotifsgenome.txt",
                converters={'pct_of_target_sequences_with_motif': lambda x: float(x.rstrip('%'))}
            ),
            {
                'motif_name': 'Top Motif', 
                'log10_p_value': 'Top P(log10)',
                'pct_of_target_sequences_with_motif': 'Top Motif Fg Pct'
            }, 
            id_col='Sample'
        ),
        left_index=True, 
        right_index=True
    )
    
    if peak_info.index.isin(motifs_of_interest.keys()).any():
        peak_info = peak_info.merge(
            parse_motif_results(
                pd.concat(find_key_motifs("homer", motifs_of_interest)),
                {
                    'motif_name': 'Key Motif', 
                    'log10_p_value': 'Key P(log10)',
                    'rank': 'Key Rank',
                    'pct_of_target_sequences_with_motif': 'Key Motif Fg Pct'
                }
            ),
            left_index=True,
            right_index=True,
            how='left'
        )
    
    peak_info = peak_info.sort_index(axis='columns')

    return df.merge(peak_info, left_on='Group', right_index=True)

def get_encode_peak_stats(peakset, filename):
    df = pd.read_table(
        filename,
        index_col='Sample',
        usecols=('Sample','Nt','Np')
    ).rename_axis(index='Group')
    return df.assign(optimal=df.max(axis=1)).rename(columns={'Nt': f"{peakset} Cons # Peaks", 'optimal': f"{peakset} Opt # Peaks"}).drop(columns='Np')

def find_key_motifs(homer_dir, motifs_of_interest):
    for group in motifs_of_interest.keys():
        pattern = '\\b' + '|'.join(motifs_of_interest[group]) + '\\b'
        for resultfile in Path(homer_dir).glob(f"{group}_*.tsv"):
            df = pd.read_table(
                resultfile,
                converters={'pct_of_target_sequences_with_motif': lambda x: float(x.rstrip('%'))}
            )
            hasmotif = df.motif_name.str.contains(pattern, case=False, na=False)
            yield df.loc[hasmotif.idxmax()].to_frame().T

def parse_motif_results(df, output_col, *, id_col='id'):
    """
    Given a DataFrame (df) containing HOMER findMotifsGenome results,
    this function:
     a) Parses out the peakset name and group id from the id_col column
     b) Creates peakset-specific columns for each column specified in
        output_col (which is a dict where the keys are columns in df
        and the values are the name of the output column -- which will
        be prepended with the peakset name)
    The resulting DataFrame is returned
    """
    df = pd.concat([df, df[id_col].str.extract("(?P<Group>.+)_(?P<Peakset>(?:idr|overlap)_(?:conservative|optimal))", expand=True)], axis=1)
    df = df.set_index(['Group', 'Peakset']).drop(columns=id_col)
    return pd.concat([
        df[col].unstack().rename(columns={
            'idr_conservative': f'IDR Cons {name}',
            'idr_optimal': f'IDR Opt {name}',
            'overlap_conservative': f'Overlap Cons {name}',
            'overlap_optimal': f'Overlap Opt {name}',
        })
        for col, name in output_col.items()
    ], axis=1)

def get_cell_coord(row, col):
    """
    Returns an Excel-style cell reference (eg "B4" or "AA315")
    for a given (row, col) index, swapping for args.bycol if necessary
    """
    if args.bycol:
        return f"{get_column_letter(row+2)}{col+2}"
    return f"{get_column_letter(col+2)}{row+2}"

def get_column(index):
    """
    Returns an Excel-style row or column reference (eg "A" or "4"),
    depending on args.bycol, for index
    """
    if args.bycol:
        return str(index+2)
    return get_column_letter(index+2)

def set_metric_display(ws, dfn, data_columns, nSamples):
    """
    Given the user-specified config (in dfn), applies number formats/etc
    to the metric rows/columns in ws
    """
    for col in dfn:
        try:
            col_idx = list(data_columns.values).index(col)
        except ValueError: # If the column isn't in the output, skip
            continue
        if 'commify' in dfn[col] or 'precision' in dfn[col]:
            numberFmt = "#,###" if dfn[col].get('commify', False) else ""
            numberFmt += "0"
            prec = dfn[col].get('precision', 0)
            if prec:
                numberFmt += "." + "0" * prec
            for cell in ws[get_column(col_idx)]:
                cell.number_format = numberFmt
        if 'colorscale' in dfn[col]:
            scaleDfn = dfn[col]['colorscale']
            rule = ColorScaleRule(
                start_type = scaleDfn.get('start', {}).get('type', None),
                start_value = scaleDfn.get('start', {}).get('val', None),
                start_color = scaleDfn.get('start', {}).get('fill', None),
                mid_type = scaleDfn.get('mid', {}).get('type', None),
                mid_value = scaleDfn.get('mid', {}).get('val', None),
                mid_color = scaleDfn.get('mid', {}).get('fill', None),
                end_type = scaleDfn.get('end', {}).get('type', None),
                end_value = scaleDfn.get('end', {}).get('val', None),
                end_color = scaleDfn.get('end', {}).get('fill', None),
            )
            ws.conditional_formatting.add(f"{get_cell_coord(0, col_idx)}:{get_cell_coord(nSamples-1, col_idx)}", rule)
        if 'highlight' in dfn[col]:
            highlightDfn = dfn[col]['highlight']
            if 'fill' in highlightDfn:
                # There is no SolidFill, use the more verbose PatternFill
                fill = PatternFill(
                    start_color=highlightDfn['fill'],
                    end_color=highlightDfn['fill'],
                    fill_type='solid'
                )
            else:
                fill = None
            rule = CellIsRule(
                operator = highlightDfn.get('operator', None),
                formula = [ highlightDfn.get('value', None) ],
                fill = fill
            )
            ws.conditional_formatting.add(f"{get_cell_coord(0, col_idx)}:{get_cell_coord(nSamples-1, col_idx)}", rule)

df = collect_experiment(Path(args.multiqc_dir), key_motifs)

text_output = f"{args.output}.txt"
excel_output = f"{args.output}.xlsx"

if args.bycol:
    df.T.to_csv(text_output, sep='\t')
    df.T.to_excel(excel_output)
else:
    df.to_csv(text_output, sep='\t')
    df.to_excel(excel_output)

wb = load_workbook(excel_output)
ws = wb.active

n_samples, n_metrics = df.shape

# Remove the borders from the row/column headers
noBorder = Border(left=Side(border_style=None), right=Side(border_style=None), top=Side(border_style=None), bottom=Side(border_style=None))
for row in range(n_samples):
    ws[get_cell_coord(row, -1)].border = noBorder
for col in range(n_metrics):
    ws[get_cell_coord(-1, col)].border = noBorder

# Make the first column the right width and display the headers at an angle
ws.column_dimensions["A"].width = 1.0 * max([len(str(cell.value)) for cell in ws["A"]])
for cell in ws["1"]:
    cell.alignment = Alignment(text_rotation=60)

# Set the remaining codes to a fixed size
shape_idx = 0 if args.bycol else 1
for col in range(df.shape[shape_idx]):
    ws.column_dimensions[get_column_letter(col+2)].width = 13

# Determine each group of samples, add a border to separate groups, and merge 
# identical cells within a group
# We can't use DataFrame.groupby here because it merges identical
# group names across analyses (ie, the control group)
if args.bycol:
    groupBorder = Border(left=Side(style='medium', color='000000'))
    centered = Alignment(horizontal='center')
else:
    groupBorder = Border(top=Side(style='medium', color='000000'))
    centered = Alignment(vertical='center')
grp_start = 0
while grp_start < n_samples:
    grp_end = grp_start + 1
    group = df.iloc[grp_start]['Group']
    while grp_end < n_samples and df.iloc[grp_end]['Group'] == group:
        grp_end += 1
    grp_end -= 1
    for i in range(-1, n_metrics):
        ws[get_cell_coord(grp_start, i)].border = groupBorder
    if grp_start != grp_end:
        for i in range(n_metrics):
            col = df.columns[i]
            if df.iloc[grp_start:grp_end+1].value_counts(col).size == 1:
                ws.merge_cells(f"{get_cell_coord(grp_start, i)}:{get_cell_coord(grp_end,i)}")
                ws[get_cell_coord(grp_start, i)].alignment = centered
    grp_start = grp_end+1

narrowLine = Side(style='thin', color='000000')
for peakSet in [f"{peak} # Peaks" for peak in ('IDR Cons', 'IDR Opt', 'Overlap Cons', 'Overlap Opt')]:
    col_idx = list(df.columns.values).index(peakSet)
    for i in range(-1, n_samples):
        cell = get_cell_coord(i, col_idx)
        oldborder = ws[cell].border
        if args.bycol:
            ws[cell].border = Border(
                left=oldborder.left,
                right=oldborder.right,
                top=narrowLine,
                bottom=oldborder.bottom
            )
        else:
            ws[cell].border = Border(
                left=narrowLine,
                right=oldborder.right,
                top=oldborder.top,
                bottom=oldborder.bottom
            )

set_metric_display(ws, metric_def["display"], df.columns, n_samples)

ws.freeze_panes = "B3" if args.bycol else "C2"

wb.save(excel_output)
